"""
app.py — Servidor Flask para el Motor Solar Fotovoltaico.
Endpoints:
  GET  /              → index.html
  POST /api/demand    → perfil de demanda (nuevos parámetros)
  POST /api/solar     → motor Jensen
  GET  /api/download  → Excel con todos los datos
"""

import io
import datetime
import traceback
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS

from demand_profile import generate_demand_profile, PLANT_PROFILES
from solar_engine import run_solar_engine

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

_cache = {}   # sesión simple en memoria

DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


# ─── Frontend ────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


# ─── Tipos de planta disponibles ─────────────────────────────────────────────
@app.route('/api/plant_types', methods=['GET'])
def api_plant_types():
    return jsonify({
        'ok': True,
        'types': [
            {'key': k, 'name': v['name'], 'desc': v['desc']}
            for k, v in PLANT_PROFILES.items()
        ]
    })


# ─── Perfil de Demanda ───────────────────────────────────────────────────────
@app.route('/api/demand', methods=['POST'])
def api_demand():
    try:
        d = request.get_json(force=True)

        Pmax             = float(d.get('pmax_kW', 50))
        FC               = float(d.get('fc_planta', 0.60))
        FP               = float(d.get('fp_potencia', 0.85))
        n_shifts         = int(d.get('n_shifts', 2))
        plant_type       = str(d.get('plant_type', 'manufactura_ligera'))
        weekend_factor   = float(d.get('weekend_op_factor', 0.50))
        summer_boost     = float(d.get('summer_boost', 1.10))

        # Validaciones
        Pmax           = max(1.0, Pmax)
        FC             = float(np.clip(FC, 0.40, 0.90))
        FP             = float(np.clip(FP, 0.60, 1.00))
        n_shifts       = int(np.clip(n_shifts, 1, 3))
        weekend_factor = float(np.clip(weekend_factor, 0.0, 1.0))
        summer_boost   = float(np.clip(summer_boost, 1.0, 1.50))

        result = generate_demand_profile(
            Pmax_kW=Pmax,
            FC_planta=FC,
            FP_potencia=FP,
            n_shifts=n_shifts,
            plant_type=plant_type,
            weekend_op_factor=weekend_factor,
            summer_boost=summer_boost,
        )

        _cache['demand'] = result

        return jsonify({
            'ok': True,
            'monthly_avg'    : result['monthly_avg'],
            'monthly_max'    : result['monthly_max'],
            'monthly_min'    : result['monthly_min'],
            'monthly_kWh'    : result['monthly_kWh'],
            'daily_profile'  : result['daily_weekday'],
            'daily_weekend'  : result['daily_weekend'],
            'stats'          : result['stats'],
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e),
                        'trace': traceback.format_exc()}), 400


# ─── Motor Solar ─────────────────────────────────────────────────────────────
@app.route('/api/solar', methods=['POST'])
def api_solar():
    try:
        d = request.get_json(force=True)

        lat         = float(np.clip(float(d.get('lat', 25.67)), -90, 90))
        lon         = float(d.get('lon', -100.31))
        alt         = float(d.get('alt', 538))
        n_panels    = max(1, int(d.get('n_panels', 50)))
        tilt        = float(np.clip(float(d.get('tilt', 25.0)), 0, 90))
        azimuth     = float(d.get('azimuth', 180.0))
        p_nominal_w = float(d.get('p_nominal_w', 400))

        result = run_solar_engine(lat, lon, alt,
                                  n_panels, tilt, azimuth, p_nominal_w)

        # Balance si hay demanda en caché
        balance = None
        if 'demand' in _cache:
            dem_arr = np.array(_cache['demand']['demand_kW'])
            gen_arr = np.array(result['P_kw_arr'])
            exceso  = np.maximum(gen_arr - dem_arr, 0)
            deficit = np.maximum(dem_arr - gen_arr, 0)
            e_dem   = float(np.sum(dem_arr) * 0.25)
            e_gen   = float(np.sum(gen_arr) * 0.25)
            cob     = min(e_gen / e_dem * 100, 100) if e_dem > 0 else 0

            monthly_balance, monthly_cob = [], []
            idx = 0
            for nd in DAYS_IN_MONTH:
                np_ = nd * 96
                eg = float(np.sum(gen_arr[idx:idx+np_]) * 0.25)
                ed = float(np.sum(dem_arr[idx:idx+np_]) * 0.25)
                monthly_balance.append(round(eg - ed, 2))
                monthly_cob.append(round(min(eg/ed*100, 100) if ed > 0 else 0, 2))
                idx += np_

            balance = {
                'energia_demanda_kWh' : round(e_dem, 2),
                'energia_generada_kWh': round(e_gen, 2),
                'cobertura_pct'       : round(cob, 2),
                'exceso_kWh'          : round(float(np.sum(exceso) * 0.25), 2),
                'deficit_kWh'         : round(float(np.sum(deficit) * 0.25), 2),
                'monthly_balance'     : monthly_balance,
                'monthly_cobertura'   : monthly_cob,
            }

        _cache['solar'] = result

        return jsonify({
            'ok'               : True,
            'monthly_gtot_avg' : result['monthly_gtot_avg'],
            'monthly_gtot_max' : result['monthly_gtot_max'],
            'monthly_gen_kWh'  : result['monthly_gen_kWh'],
            'daily_gtot_summer': result['daily_gtot_summer'],
            'daily_p_summer'   : result['daily_p_summer'],
            'stats'            : result['stats'],
            'balance'          : balance,
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e),
                        'trace': traceback.format_exc()}), 400


# ─── Descarga Excel ──────────────────────────────────────────────────────────
@app.route('/api/download/excel', methods=['GET'])
def api_download_excel():
    if 'solar' not in _cache:
        return jsonify({'ok': False,
                        'error': 'Ejecuta primero el Motor Solar.'}), 400

    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference

    solar  = _cache['solar']
    demand = _cache.get('demand')

    hours  = solar['hours']
    Gtot   = solar['Gtot_arr']
    P_kw   = solar['P_kw_arr']
    dem_kw = demand['demand_kW'] if demand else [0] * len(hours)

    base_dt = datetime.datetime(2024, 1, 1, 0, 0)
    fechas  = [(base_dt + datetime.timedelta(hours=h)).strftime('%Y-%m-%d %H:%M')
               for h in hours]

    wb = openpyxl.Workbook()

    # ── Estilos comunes ──
    HDR_FILL   = PatternFill("solid", fgColor="0D1526")
    COL_FILLS  = {
        'fecha'  : PatternFill("solid", fgColor="111827"),
        'irrad'  : PatternFill("solid", fgColor="1a1f0a"),
        'gen'    : PatternFill("solid", fgColor="0a1a14"),
        'dem'    : PatternFill("solid", fgColor="1a100a"),
        'bal'    : PatternFill("solid", fgColor="0a0a1a"),
    }
    HDR_FONT   = Font(name='Calibri', bold=True, color='F97316', size=10)
    DATA_FONT  = Font(name='Calibri', size=9)
    TITLE_FONT = Font(name='Calibri', bold=True, color='FBBF24', size=13)
    KPI_FONT   = Font(name='Calibri', bold=True, color='F1F5F9', size=11)
    thin       = Side(style='thin', color='1E293B')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(ws, row, col, val, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font  = HDR_FONT
        c.fill  = fill or HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)
        c.border = border
        return c

    def data_cell(ws, row, col, val, fmt=None, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = DATA_FONT
        c.border = border
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 1 — Parámetros
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = '1. Parámetros'
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 22

    s = solar['stats']
    title = ws1.cell(row=1, column=1,
                     value='Motor Solar FV — Parámetros de Simulación')
    title.font = TITLE_FONT
    ws1.merge_cells('A1:B1')

    params = [
        ('── SISTEMA PV ──', ''),
        ('Latitud',            f"{s['lat']} °"),
        ('Longitud',           f"{s['lon']} °"),
        ('Altitud',            f"{s['alt']} m s.n.m."),
        ('Potencia nominal',   f"{s['potencia_nominal_W_panel']} W"),
        ('Número de paneles',  s['n_paneles']),
        ('Inclinación (tilt)', f"{s['tilt']} °"),
        ('Azimut',             f"{s['azimuth']} °"),
        ('', ''),
    ]
    if demand:
        ds = demand['stats']
        params += [
            ('── PLANTA INDUSTRIAL ──', ''),
            ('Tipo de planta',      ds.get('plant_name', '—')),
            ('Demanda máxima',      f"{ds['pmax_kW']:.1f} kW"),
            ('Turnos de operación', ds.get('n_shifts', '—')),
            ('Factor de carga',     f"{ds['FC_planta']*100:.0f} %"),
            ('Factor de potencia',  f"{ds['FP_potencia']:.2f}"),
            ('Op. fin de semana',   f"{ds.get('weekend_op_factor',0)*100:.0f} %"),
            ('Boost verano',        f"×{ds.get('summer_boost',1):.2f}"),
        ]

    for i, (k, v) in enumerate(params, start=3):
        c_k = ws1.cell(row=i, column=1, value=k)
        c_v = ws1.cell(row=i, column=2, value=v)
        if k.startswith('──'):
            c_k.font = Font(name='Calibri', bold=True, color='F97316', size=10)
            c_v.value = ''
        else:
            c_k.font  = DATA_FONT
            c_v.font  = Font(name='Calibri', bold=True, color='F1F5F9', size=9)
        for c in [c_k, c_v]:
            c.border    = border
            c.alignment = Alignment(vertical='center')

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 2 — KPIs / Resumen Ejecutivo
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('2. KPIs')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 14

    t2 = ws2.cell(row=1, column=1, value='Resumen Ejecutivo — KPIs del Sistema')
    t2.font = TITLE_FONT
    ws2.merge_cells('A1:C1')

    kpis_solar = [
        ('☀️  GENERACIÓN SOLAR', '', ''),
        ('Energía anual generada', f"{s['energia_anual_kWh']:,.0f}", 'kWh/año'),
        ('Energía anual generada', f"{s['energia_anual_MWh']:,.2f}", 'MWh/año'),
        ('Potencia pico del sistema', f"{s['p_nominal_total_kW']:,.2f}", 'kWp'),
        ('Factor de capacidad', f"{s['factor_capacidad_pct']:,.2f}", '%'),
        ('Horas pico solar equiv.', f"{s['horas_pico_sol_equiv']:,.0f}", 'hrs/año'),
        ('Horas con generación > 0', f"{s['n_horas_generacion']:,.0f}", 'hrs/año'),
        ('Irradiación horizontal', f"{s['irrad_horizontal_kWh_m2']:,.0f}", 'kWh/m²·año'),
        ('Irradiación POA total', f"{s['irrad_poa_kWh_m2']:,.0f}", 'kWh/m²·año'),
        ('Irradiancia POA máxima', f"{s['gtot_max_W_m2']:,.1f}", 'W/m²'),
        ('Irradiancia POA media (días sol.)', f"{s['gtot_media_W_m2']:,.1f}", 'W/m²'),
    ]

    if demand:
        ds = demand['stats']
        kpis_solar += [
            ('', '', ''),
            ('🏭  DEMANDA INDUSTRIAL', '', ''),
            ('Energía demandada anual', f"{ds['energia_anual_kWh']:,.0f}", 'kWh/año'),
            ('Demanda media', f"{ds['p_media_kW']:,.2f}", 'kW'),
            ('Demanda máxima real', f"{ds['p_max_real_kW']:,.2f}", 'kW'),
            ('Factor de carga real', f"{ds['factor_carga_real']*100:,.1f}", '%'),
            ('Horas equiv. plena carga', f"{ds['horas_punta_equiv']:,.0f}", 'hrs/año'),
        ]

    if 'balance' in (solar if False else {}):
        pass  # balance se calcula en caché aparte

    # Calcular balance aquí para KPIs
    if demand:
        dem_arr = np.array(demand['demand_kW'])
        gen_arr = np.array(P_kw)
        e_dem   = float(np.sum(dem_arr) * 0.25)
        e_gen   = float(np.sum(gen_arr) * 0.25)
        cob     = min(e_gen / e_dem * 100, 100) if e_dem > 0 else 0
        exceso  = float(np.sum(np.maximum(gen_arr - dem_arr, 0)) * 0.25)
        deficit = float(np.sum(np.maximum(dem_arr - gen_arr, 0)) * 0.25)
        kpis_solar += [
            ('', '', ''),
            ('⚡  BALANCE ENERGÉTICO', '', ''),
            ('Cobertura solar de demanda', f"{cob:.1f}", '%'),
            ('Energía cubierta por solar', f"{min(e_gen,e_dem):,.0f}", 'kWh/año'),
            ('Excedente solar (inyección red)', f"{exceso:,.0f}", 'kWh/año'),
            ('Déficit (energía de la red)', f"{deficit:,.0f}", 'kWh/año'),
        ]

    hdr(ws2, 2, 1, 'Indicador')
    hdr(ws2, 2, 2, 'Valor')
    hdr(ws2, 2, 3, 'Unidad')

    for i, (k, v, u) in enumerate(kpis_solar, start=3):
        if k.endswith('──') or k.endswith('SOLAR') or k.endswith('INDUSTRIAL') or k.endswith('ENERGÉTICO'):
            c = ws2.cell(row=i, column=1, value=k)
            c.font = Font(name='Calibri', bold=True, color='F97316', size=10)
            c.fill = PatternFill("solid", fgColor="0D1526")
            ws2.merge_cells(f'A{i}:C{i}')
            c.border = border
        else:
            data_cell(ws2, i, 1, k)
            c = data_cell(ws2, i, 2, v)
            c.alignment = Alignment(horizontal='right')
            c.font = Font(name='Calibri', bold=True, color='FBBF24', size=9)
            data_cell(ws2, i, 3, u)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 3 — Resumen Mensual
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('3. Resumen Mensual')
    ws3.sheet_view.showGridLines = False
    MONTHS_ES = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                  'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    t3 = ws3.cell(row=1, column=1, value='Resumen Mensual — Irradiancia, Generación y Demanda')
    t3.font = TITLE_FONT
    ws3.merge_cells('A1:H1')

    headers3 = ['Mes', 'Irrad. POA Avg\n[W/m²]', 'Irrad. POA Máx\n[W/m²]',
                 'Generación\n[kWh/mes]', 'Generación\n[MWh/mes]',
                 'Demanda\n[kWh/mes]', 'Demanda\n[MWh/mes]', 'Cobertura\n[%]']
    for ci, h in enumerate(headers3, 1):
        hdr(ws3, 2, ci, h)
        ws3.column_dimensions[get_column_letter(ci)].width = 16

    # Calcular demanda mensual
    if demand:
        dem_arr  = np.array(demand['demand_kW'])
        dem_mon  = []
        idx = 0
        for nd in DAYS_IN_MONTH:
            seg = dem_arr[idx:idx+nd*96]
            dem_mon.append(float(np.sum(seg) * 0.25))
            idx += nd * 96
    else:
        dem_mon = [0] * 12

    for mi, mes in enumerate(MONTHS_ES):
        r   = mi + 3
        gen = solar['monthly_gen_kWh'][mi]
        dem = dem_mon[mi]
        cob = min(gen / dem * 100, 100) if dem > 0 else 0
        row_data = [
            mes,
            round(solar['monthly_gtot_avg'][mi], 1),
            round(solar['monthly_gtot_max'][mi], 1),
            round(gen, 0),
            round(gen / 1000, 3),
            round(dem, 0),
            round(dem / 1000, 3),
            round(cob, 1),
        ]
        fills3 = [None, COL_FILLS['irrad'], COL_FILLS['irrad'],
                  COL_FILLS['gen'],  COL_FILLS['gen'],
                  COL_FILLS['dem'],  COL_FILLS['dem'],
                  COL_FILLS['bal']]
        fmts3  = [None, '#,##0.0', '#,##0.0',
                  '#,##0', '#,##0.000',
                  '#,##0', '#,##0.000', '0.0"%"']
        for ci, (val, fill, fmt) in enumerate(zip(row_data, fills3, fmts3), 1):
            c = data_cell(ws3, r, ci, val, fmt=fmt, fill=fill)
            if ci == 1:
                c.font = Font(name='Calibri', bold=True, color='F1F5F9', size=9)

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 4 — Datos Anuales (35,040 filas)
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('4. Datos Anuales (15-min)')
    ws4.sheet_view.showGridLines = False

    headers4 = ['Fecha-Hora', 'Irradiancia POA\n[W/m²]',
                 'Generación PV\n[kW]', 'Demanda\n[kW]',
                 'Balance\n[kW]']
    col_widths4 = [18, 18, 15, 15, 15]
    for ci, (h, w) in enumerate(zip(headers4, col_widths4), 1):
        hdr(ws4, 1, ci, h)
        ws4.column_dimensions[get_column_letter(ci)].width = w

    fills4 = [COL_FILLS['fecha'], COL_FILLS['irrad'],
              COL_FILLS['gen'],   COL_FILLS['dem'],   COL_FILLS['bal']]
    fmts4  = [None, '#,##0.0', '#,##0.00', '#,##0.00', '#,##0.00']

    dem_kw_arr = dem_kw if demand else [0] * len(hours)

    for i, (f, g, p, dm) in enumerate(zip(fechas, Gtot, P_kw, dem_kw_arr), start=2):
        bal = p - dm
        row_vals = [f, round(g, 2), round(p, 4), round(dm, 4), round(bal, 4)]
        for ci, (val, fill, fmt) in enumerate(zip(row_vals, fills4, fmts4), 1):
            data_cell(ws4, i, ci, val, fmt=fmt, fill=fill)

    # Congelar primera fila de datos
    ws4.freeze_panes = 'A2'

    # ══════════════════════════════════════════════════════════════════════════
    # HOJA 5 — Perfil Diario
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('5. Perfil Diario')
    ws5.sheet_view.showGridLines = False

    t5 = ws5.cell(row=1, column=1, value='Perfil Diario Promedio — Generación vs. Demanda')
    t5.font = TITLE_FONT
    ws5.merge_cells('A1:E1')

    hours_96 = [f"{h//4:02d}:{(h%4)*15:02d}" for h in range(96)]
    headers5 = ['Hora', 'Gen. PV Verano\n[kW]', 'Dem. Laboral\n[kW]',
                'Dem. Fin de Semana\n[kW]', 'Balance Laboral\n[kW]']
    for ci, h in enumerate(headers5, 1):
        hdr(ws5, 2, ci, h)
        ws5.column_dimensions[get_column_letter(ci)].width = 18

    gp_summer  = solar['daily_p_summer']
    dw_weekday = demand['daily_weekday'] if demand else [0]*96
    dw_weekend = demand['daily_weekend'] if demand else [0]*96

    for i, (hr, gp, dwd, dwe) in enumerate(
            zip(hours_96, gp_summer, dw_weekday, dw_weekend), start=3):
        bal = gp - dwd
        ws5.cell(row=i, column=1, value=hr).font = DATA_FONT
        for ci, val in enumerate([gp, dwd, dwe, bal], start=2):
            c = data_cell(ws5, i, ci, round(val, 3), fmt='#,##0.000')
            if ci == 5:  # balance: verde si >0, rojo si <0
                c.font = Font(name='Calibri', size=9,
                               color='10B981' if val >= 0 else 'EF4444', bold=True)

    # ── Guardar y enviar ──────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Motor_Solar_FV_Resultados.xlsx'
    )


# ─── CSV legacy (por compatibilidad) ─────────────────────────────────────────
@app.route('/api/download', methods=['GET'])
def api_download_csv():
    return api_download_excel()


# ─── Arranque ─────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("  Motor Solar Fotovoltaico — Servidor Flask")
    print("  Abrir en navegador: http://localhost:5000")
    print("=" * 60)
    app.run(debug=True, port=5000, use_reloader=False)
